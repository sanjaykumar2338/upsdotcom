import csv, os, sys, time, random, pathlib, threading, queue, re, base64
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

"""
Priority UPS Script (Order #1)
 - Mirrors threading + file-locking style of run_option_two_concurrent.py
 - Accepts CSV/XLSX input of ZIP codes
 - Outputs CSV with city/state/zip, ship date, image URL, and optional downloaded image path
 - GUI provides Select Input/Output + Start/Pause/Resume + status counter
 - Anti-blocking: pauses + clears cookies on HTTP failures and every 250 queries
"""

# ---------- Editable toggles ----------
SAVE_IMAGES      = False          # download map images (set False for data-only runs)
MAX_WORKERS      = 2             # max worker threads
HTTP_TIMEOUT     = 60            # per-request timeout (seconds)
MAX_RETRIES      = 3             # retries per ZIP before SKIPPED
IMAGE_RETRIES    = 3             # retries per image download
UNRESPONSIVE_PAUSE = 120         # seconds to pause after HTTP errors before retrying
MAINTENANCE_EVERY  = 250         # after this many processed queries: pause + clear cookies
MAINTENANCE_PAUSE  = 5           # seconds to sleep during maintenance
USER_AGENT       = "Mozilla/5.0 (PriorityUPS/1.0)"
BASE_URL         = "https://www.ups.com/maps/?loc=en_CB"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}
HUMAN_DELAY_RANGE_MS = (1000, 2000)  # shorter human-like delay after UPS requests (1-2 seconds)
DEBUG_SCREENSHOT_DIR = "debug_shots"
HEADLESS = False                  # keep False; browser is minimized immediately to stay out of the way
os.environ.setdefault("SE_MANAGER_TELEMETRY", "0")

# ---------- Thread-safe state ----------
write_lock = threading.Lock()
stats_lock = threading.Lock()
session_version_lock = threading.Lock()
maintenance_lock = threading.Lock()
log_lock = threading.Lock()
browser_lock = threading.Lock()
drivers = []  # track created browser instances for cleanup

thread_local = threading.local()
pause_event = threading.Event()
pause_event.set()
stop_event = threading.Event()

stats = {"processed": 0, "skipped": 0, "total": 0, "last": ""}
session_version = 0
processed_since_maintenance = 0

# ---------- Helpers ----------
def _space(s): return " ".join((s or "").split())


def log(msg):
    with log_lock:
        print(msg, flush=True)


def bump_session_version():
    global session_version
    with session_version_lock:
        session_version += 1


def new_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def get_session():
    ver = session_version
    if not hasattr(thread_local, "sess") or getattr(thread_local, "sess_version", None) != ver:
        thread_local.sess = new_session()
        thread_local.sess_version = ver
    return thread_local.sess


def fetch_image_via_browser(driver, url):
    """
    Fetch image bytes using the live browser context (respects UPS cookies).
    Returns raw bytes or raises on failure.
    """
    script = """
    const url = arguments[0];
    const done = arguments[arguments.length - 1];
    fetch(url, {credentials: 'include'})
      .then(res => res.arrayBuffer())
      .then(buf => {
        const bytes = Array.from(new Uint8Array(buf));
        done(btoa(String.fromCharCode.apply(null, bytes)));
      })
      .catch(err => done({error: String(err)}));
    """
    res = driver.execute_async_script(script, url)
    if isinstance(res, dict) and res.get("error"):
        raise RuntimeError(res["error"])
    try:
        return base64.b64decode(res)
    except Exception as e:
        raise RuntimeError(f"decode_error:{e}")


def normalize_zip(val: str) -> str:
    v = (val or "").strip()
    if v.isdigit() and len(v) < 5:
        v = v.zfill(5)
    return v


def is_internet_available(url: str = "https://www.google.com/generate_204", timeout: int = 4) -> bool:
    try:
        requests.get(url, timeout=timeout)
        return True
    except Exception:
        return False


def _detect_zip_index(headers):
    ZIP_HEADER_ALIASES = {"zip", "zipcode", "zip_code", "zipcodes", "postal", "postalcode"}
    for idx, h in enumerate(headers):
        if not h:
            continue
        header_lower = str(h).strip().lower()
        if header_lower in ZIP_HEADER_ALIASES:
            return idx
    return 0


def read_input_rows(path: str):
    """
    Read ZIP rows from xlsx/csv/txt, preserving leading zeros where possible.
    """
    zips = []
    p = pathlib.Path(path)
    ext = p.suffix.lower()

    if ext == ".xlsx":
        if not openpyxl:
            raise RuntimeError("openpyxl not installed; required for XLSX input.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return zips
        headers = [c if c is not None else "" for c in rows[0]]
        zip_idx = _detect_zip_index(headers)
        data_rows = rows[1:] if any(headers) else rows
        for row in data_rows:
            if not row:
                continue
            val = row[zip_idx] if zip_idx < len(row) else None
            if val is None:
                continue
            if isinstance(val, (int, float)):
                val = str(int(val))
            z = normalize_zip(str(val))
            if z:
                zips.append(z)
        wb.close()
    elif ext == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return zips
        headers = rows[0]
        zip_idx = _detect_zip_index(headers)
        data_rows = rows[1:] if any(headers) else rows
        for row in data_rows:
            if not row:
                continue
            val = row[zip_idx] if zip_idx < len(row) else ""
            z = normalize_zip(str(val))
            if z:
                zips.append(z)
    elif ext == ".txt":
        with open(path, encoding="utf-8-sig") as f:
            for line in f:
                z = normalize_zip(line.strip())
                if z:
                    zips.append(z)
    else:
        raise RuntimeError("Unsupported input type. Use .xlsx, .csv, or .txt")
    return zips


def get_browser():
    """
    Return a Selenium Chrome driver for this thread. Creates if missing.
    Runs visible (not headless) to satisfy UPS, then minimizes to stay unobtrusive.
    """
    if hasattr(thread_local, "driver") and thread_local.driver:
        return thread_local.driver
    try:
        from selenium import webdriver
    except ImportError as e:
        raise RuntimeError("selenium not installed. Install with: py -m pip install selenium") from e

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(options=options)
    with browser_lock:
        drivers.append(driver)
    try:
        driver.minimize_window()
    except Exception:
        pass
    driver.set_page_load_timeout(HTTP_TIMEOUT)
    thread_local.driver = driver
    return driver


def close_all_browsers():
    with browser_lock:
        active = list(drivers)
        drivers.clear()
    for d in active:
        try:
            d.quit()
        except Exception:
            pass


def ensure_header(path):
    cols = [
        "ZIP",
        "CITY",
        "STATE",
        "SHIP_DATE",
        "LOCATION_TEXT",
        "IMAGE_URL",
        "IMAGE_FILE",
        "STATUS",
        "ERROR_STEP",
        "ERROR_TYPE",
        "ERROR_MESSAGE",
        "RAW_URL",
    ]

    if (not os.path.exists(path)) or os.path.getsize(path) == 0:
        with open(path, "w", newline="", encoding="utf-8") as fp:
            csv.writer(fp).writerow(cols)
        return cols

    try:
        with open(path, newline="", encoding="utf-8") as fp:
            reader = csv.reader(fp)
            first_row = next(reader, None)
            if first_row != cols:
                rows = [first_row] + list(reader) if first_row else list(reader)
                with open(path, "w", newline="", encoding="utf-8") as out:
                    w = csv.writer(out)
                    w.writerow(cols)
                    w.writerows(r for r in rows if r)
    except Exception:
        with open(path, "w", newline="", encoding="utf-8") as fp:
            csv.writer(fp).writerow(cols)
    return cols


def append_row(path, cols, row, tries=20, base_sleep=0.25):
    for i in range(tries):
        try:
            with write_lock, open(path, "a", newline="", encoding="utf-8") as fp:
                w = csv.DictWriter(fp, fieldnames=cols)
                w.writerow(row)
                fp.flush()
                os.fsync(fp.fileno())
            return
        except PermissionError:
            time.sleep(base_sleep * (1 + i / 3.0))
    alt = f"{os.path.splitext(path)[0]}_{os.getpid()}_{int(time.time())}.csv"
    with write_lock, open(alt, "a", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=cols)
        if fp.tell() == 0:
            w.writeheader()
        w.writerow(row)


def make_error(error_step, error_type, message):
    error_step = (error_step or "").strip()
    error_type = (error_type or "").strip()
    message = (message or "").strip()
    return error_step, error_type, message[:250]


def load_processed_zips(path):
    processed = set()
    if not os.path.exists(path):
        return processed
    try:
        with open(path, newline="", encoding="utf-8") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                z = (row.get("ZIP") or "").strip()
                if z:
                    processed.add(z)
    except Exception:
        pass
    return processed


def wait_if_paused():
    while not pause_event.is_set():
        if stop_event.is_set():
            return False
        time.sleep(0.1)
    return not stop_event.is_set()


def load_zips(path):
    path = pathlib.Path(path)
    zips = []
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            first = True
            for row in csv.reader(f):
                if not row:
                    continue
                v = str(row[0]).strip()
                if first and not v.isdigit():
                    first = False
                    continue
                first = False
                if v:
                    zips.append(v)
    else:
        if not openpyxl:
            raise RuntimeError("openpyxl not installed; required for XLSX input.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        first = True
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            v = str(row[0]).strip() if row[0] is not None else ""
            if first and not v.isdigit():
                first = False
                continue
            first = False
            if v:
                zips.append(v)
        wb.close()
    return zips


def maybe_maintenance():
    global processed_since_maintenance
    with maintenance_lock:
        if processed_since_maintenance >= MAINTENANCE_EVERY:
            processed_since_maintenance = 0
            bump_session_version()
            time.sleep(MAINTENANCE_PAUSE)
            return True
    return False


def record_processed(skipped=False):
    global processed_since_maintenance
    with stats_lock:
        stats["processed"] += 1
        if skipped:
            stats["skipped"] += 1
        processed_since_maintenance += 1


def parse_results(html):
    soup = BeautifulSoup(html, "html.parser")
    text = _space(soup.get_text(" ", strip=True))

    city = state = zip_code = ship_date = ""
    loc_text = ""
    m = re.search(
        r"Business days in transit\s+(\d{1,2}/\d{1,2}/\d{4})\s+from:\s+([A-Za-z .'-]+),\s*([A-Z]{2})\s*(\d{5})",
        text,
        re.I,
    )
    if m:
        ship_date = m.group(1)
        city = m.group(2).strip()
        state = m.group(3)
        zip_code = m.group(4)
        loc_text = f"{city}, {state} {zip_code}"
    else:
        m = re.search(r"([A-Z][A-Za-z .'-]+),\s*([A-Z]{2})\s*(\d{5})", text)
        if m:
            city = m.group(1).strip()
            state = m.group(2)
            zip_code = m.group(3)
            loc_text = f"{city}, {state} {zip_code}"

    img = soup.find("img", id="imgMap") or soup.find("img", attrs={"alt": re.compile("Time in Transit Map", re.I)})
    img_url = ""
    if img and img.get("src"):
        img_url = img["src"]
    return {
        "city": city,
        "state": state,
        "zip": zip_code,
        "ship_date": ship_date,
        "location_text": loc_text,
        "image_url": img_url,
        "raw_text": text,
    }


def submit_zip(zipcode):
    sess = get_session()
    # Initial page to capture form + cookies
    try:
        base_resp = sess.get(BASE_URL, headers=HEADERS, timeout=HTTP_TIMEOUT)
        base_resp.raise_for_status()
    except requests.RequestException as e:
        return None, f"exception:{str(e)[:200]}"

    soup = BeautifulSoup(base_resp.text, "html.parser")
    form = soup.find("form")
    data = {}
    method = "get"
    action = BASE_URL

    if form:
        method = (form.get("method") or "get").lower()
        action = urljoin(BASE_URL, form.get("action") or BASE_URL)
        for inp in form.find_all("input"):
            name = inp.get("name")
            if not name:
                continue
            typ = (inp.get("type") or "text").lower()
            val = inp.get("value") or ""
            if typ in ("radio", "checkbox"):
                if inp.has_attr("checked"):
                    data[name] = val or "on"
            else:
                data[name] = val
        # best-guess zip field
        zip_field = None
        for k in data.keys():
            if "zip" in k.lower() or "postal" in k.lower():
                zip_field = k
                break
        if not zip_field:
            zip_field = "postalCode"
        data[zip_field] = zipcode
    else:
        # Fallback: hope GET parameters work
        data = {"loc": "en_US", "postalCode": zipcode}
        action = BASE_URL
        method = "get"

    try:
        if method == "post":
            resp = sess.post(action, data=data, headers=HEADERS, timeout=HTTP_TIMEOUT)
        else:
            resp = sess.get(action, params=data, headers=HEADERS, timeout=HTTP_TIMEOUT)
        resp.raise_for_status()
    except requests.RequestException as e:
        return None, f"exception:{str(e)[:200]}"
    return resp, ""


def download_image(session, img_url, image_dir, zipcode):
    if not img_url:
        return "", "no_image_url"
    full_url = urljoin(BASE_URL, img_url)
    name = os.path.basename(full_url.split("?")[0]) or f"{zipcode}.png"
    dest = image_dir / name
    if dest.exists():
        return str(dest), ""
    try:
        r = session.get(full_url, timeout=HTTP_TIMEOUT)
        if not r.ok:
            return "", f"http_img_{r.status_code}"
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as fp:
            fp.write(r.content)
        return str(dest), ""
    except requests.RequestException as e:
        return "", f"exception:{str(e)[:200]}"


# ---------- Debug: headful browser fetch ----------
def debug_headful(zipcode: str) -> Optional[str]:
    """
    Open a visible browser to observe what UPS returns.
    Saves a screenshot and HTML for manual inspection.
    Returns path to screenshot or None on failure.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
    except ImportError:
        log("Selenium not installed. Install with: py -m pip install selenium")
        return None

    zipcode = str(zipcode).strip()
    shot_dir = pathlib.Path(DEBUG_SCREENSHOT_DIR)
    shot_dir.mkdir(parents=True, exist_ok=True)
    shot_path = shot_dir / f"debug_{zipcode}.png"
    html_path = shot_dir / f"debug_{zipcode}.html"

    options = webdriver.ChromeOptions()
    # Headful by default; you can comment the next lines if Chrome auto-closure is an issue.
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(options=options)
    try:
        driver.get(BASE_URL)
        wait = WebDriverWait(driver, 20)
        # Try to find a text input to enter ZIP
        try:
            field = wait.until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//input[@type='text' or @type='search' or contains(@name,'zip') or contains(@name,'postal')]",
                    )
                )
            )
            field.clear()
            field.send_keys(zipcode)
        except Exception:
            log(f"[DEBUG] Could not find ZIP input automatically for {zipcode}")

        # Try to click submit
        try:
            btn = driver.find_element(By.XPATH, "//button[contains(., 'Submit')] | //input[@type='submit']")
            btn.click()
        except Exception:
            log("[DEBUG] Could not find Submit button; waiting for any navigation")

        time.sleep(8)  # allow page to load
        driver.save_screenshot(str(shot_path))
        with open(html_path, "w", encoding="utf-8") as fp:
            fp.write(driver.page_source)
        log(f"[DEBUG] Saved screenshot to {shot_path} and HTML to {html_path}")
        return str(shot_path)
    finally:
        driver.quit()


def process_zip(zipcode, cols, out_path, image_dir):
    zipcode_raw = str(zipcode).strip()
    zipcode = normalize_zip(zipcode_raw)
    with stats_lock:
        stats["last"] = zipcode
    log(f"[START] ZIP {zipcode}")

    parsed = {"city": "", "state": "", "zip": "", "ship_date": "", "location_text": "", "image_url": ""}
    img_path = ""
    error_message = ""
    error_type = ""
    error_step = ""
    status_val = "SKIPPED"

    try:
        attempts = 0
        while attempts < MAX_RETRIES and not stop_event.is_set():
            if not wait_if_paused():
                break

            if not is_internet_available():
                error_step, error_type, error_message = make_error(
                    "OFFLINE", "NO_INTERNET", "Internet not available; paused and retrying."
                )
                log("[NET] No internet connection detected. Pausing 120s.")
                time.sleep(UNRESPONSIVE_PAUSE)
                continue

            maybe_maintenance()
            attempts += 1

            try:
                from selenium.common.exceptions import TimeoutException, WebDriverException
                from selenium.webdriver.common.by import By
                from selenium.webdriver.common.keys import Keys
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC

                driver = get_browser()
                wait = WebDriverWait(driver, 10)

                try:
                    driver.get(BASE_URL)
                except WebDriverException:
                    error_step, error_type, error_message = make_error(
                        "LOAD_PAGE", "TIMEOUT", "Timeout / Site not responding"
                    )
                    bump_session_version()
                    continue

                try:
                    field = wait.until(
                        EC.presence_of_element_located(
                            (
                                By.XPATH,
                                "//input[@type='text' or @type='search' or contains(@name,'zip') or contains(@name,'postal')]",
                            )
                        )
                    )
                    field.clear()
                    field.send_keys(zipcode)
                except Exception:
                    error_step, error_type, error_message = make_error(
                        "FIND_INPUT",
                        "CAPTCHA_OR_BLOCKED",
                        "ZIP input not found (captcha/block/DOM changed).",
                    )
                    bump_session_version()
                    continue

                try:
                    btn = wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "//button[@type='submit' or contains(., 'Submit')] | //input[@type='submit']")
                        )
                    )
                    btn.click()
                except Exception:
                    try:
                        field.send_keys(Keys.ENTER)
                    except Exception:
                        pass

                def result_or_error(d):
                    try:
                        if d.find_elements(By.CSS_SELECTOR, "img#imgMap, img[src*='servicemaps']"):
                            return True
                        ps = (d.page_source or "").lower()
                        if ("no information for zip code" in ps) or ("either the zip code does not exist" in ps):
                            return True
                        return False
                    except Exception:
                        return False

                try:
                    wait.until(result_or_error)
                except TimeoutException:
                    error_step, error_type, error_message = make_error(
                        "WAIT_RESULT", "RESULT_TIMEOUT", "Timed out waiting for results/map image."
                    )
                    bump_session_version()
                    continue

                html = driver.page_source
                parsed = parse_results(html)
                if not parsed["location_text"]:
                    lower_ps = html.lower()
                    if ("no information for zip code" in lower_ps) or ("either the zip code does not exist" in lower_ps):
                        error_step, error_type, error_message = make_error(
                            "WAIT_RESULT", "NO_DATA", "No information for ZIP code"
                        )
                        status_val = "SKIPPED"
                        break
                    error_step, error_type, error_message = make_error(
                        "PARSE", "PARSE_FAILED", "Result loaded but expected fields not found."
                    )
                    bump_session_version()
                    continue

                img_url = parsed["image_url"]
                img_error = ""
                if img_url and SAVE_IMAGES:
                    try:
                        full_img_url = urljoin(BASE_URL, img_url)

                        session = requests.Session()
                        for cookie in driver.get_cookies():
                            session.cookies.set(cookie["name"], cookie["value"])
                        try:
                            user_agent = driver.execute_script("return navigator.userAgent;")
                        except Exception:
                            user_agent = ""
                        if user_agent:
                            session.headers.update({"User-Agent": user_agent})
                        session.headers.update({"Referer": BASE_URL})

                        base_name = os.path.basename(full_img_url.split("?")[0])
                        ext = pathlib.Path(base_name).suffix or ""
                        img_name = f"{zipcode}{ext or '.png'}"
                        dest = image_dir / img_name
                        dest.parent.mkdir(parents=True, exist_ok=True)

                        last_exc = None
                        got_bytes = None
                        for attempt in range(1, IMAGE_RETRIES + 1):
                            try:
                                img_resp = session.get(full_img_url, timeout=HTTP_TIMEOUT, allow_redirects=True)
                                img_resp.raise_for_status()
                                got_bytes = img_resp.content
                                if not ext:
                                    ctype = img_resp.headers.get("Content-Type", "")
                                    if "png" in ctype:
                                        ext = ".png"
                                    elif "gif" in ctype:
                                        ext = ".gif"
                                    elif "jpeg" in ctype or "jpg" in ctype:
                                        ext = ".jpg"
                                    img_name = f"{zipcode}{ext or '.png'}"
                                    dest = image_dir / img_name
                                last_exc = None
                                break
                            except Exception as e:
                                last_exc = e
                                log(f"[IMG-RETRY] ZIP {zipcode} attempt {attempt}/{IMAGE_RETRIES} failed: {e}")
                                time.sleep(1.5 * attempt)
                        if got_bytes is None and last_exc:
                            try:
                                got_bytes = fetch_image_via_browser(driver, full_img_url)
                                last_exc = None
                            except Exception as e:
                                last_exc = e
                        if got_bytes is None and last_exc:
                            raise last_exc
                        with open(dest, "wb") as fp:
                            fp.write(got_bytes)
                        rel_path = os.path.relpath(dest, pathlib.Path(out_path).parent)
                        img_path = rel_path
                        log(f"[IMG] Saved {img_path}")
                    except Exception as e:
                        img_error = f"Image download failed: {e}"
                        error_step, error_type, error_message = make_error(
                            "DOWNLOAD_IMAGE", "IMAGE_TIMEOUT", img_error
                        )
                        log(f"[IMG-ERR] ZIP {zipcode} | {img_error}")

                if img_error:
                    status_val = "OK_WITH_IMAGE_ERROR"
                else:
                    status_val = "OK"
                break
            finally:
                sleep_time = random.uniform(HUMAN_DELAY_RANGE_MS[0], HUMAN_DELAY_RANGE_MS[1]) / 1000.0
                log(f"[WAIT] {sleep_time:.2f}s before next request")
                time.sleep(sleep_time)

        if status_val not in ("OK", "OK_WITH_IMAGE_ERROR"):
            status_val = "SKIPPED"
            if not error_message:
                error_step, error_type, error_message = make_error("UNKNOWN", "UNKNOWN", "Timeout / Site not responding")
    except Exception as e:
        status_val = "ERROR"
        error_step, error_type, error_message = make_error("UNEXPECTED", "EXCEPTION", f"Unexpected error: {e}")
        log(f"[ERROR] ZIP {zipcode}: {e}")

    row = {
        "ZIP": zipcode,
        "CITY": parsed.get("city", ""),
        "STATE": parsed.get("state", ""),
        "SHIP_DATE": parsed.get("ship_date", ""),
        "LOCATION_TEXT": parsed.get("location_text", ""),
        "IMAGE_URL": urljoin(BASE_URL, parsed.get("image_url", "")) if parsed.get("image_url") else "",
        "IMAGE_FILE": img_path,
        "STATUS": status_val,
        "ERROR_STEP": error_step,
        "ERROR_TYPE": error_type,
        "ERROR_MESSAGE": error_message or "",
        "RAW_URL": BASE_URL,
    }
    append_row(out_path, cols, row)
    record_processed(skipped=(status_val != "OK"))
    log(f"[{status_val}] ZIP {zipcode} | location={parsed.get('location_text','')} | img_url={'yes' if parsed.get('image_url') else 'no'} | err={error_message}")


def run_batch(input_path, output_path):
    try:
        import selenium  # noqa: F401
    except ImportError:
        raise RuntimeError("selenium not installed. Install with: py -m pip install selenium")
    zips = read_input_rows(input_path)
    if not zips:
        raise RuntimeError("No ZIP codes found in input.")
    out_path = str(pathlib.Path(output_path).resolve())
    cols = ensure_header(out_path)
    already = load_processed_zips(out_path)
    if already:
        zips = [z for z in zips if z not in already]
        log(f"[JOB] Skipping {len(already)} already-processed ZIPs from existing output.")
    if not zips:
        log("[JOB] Nothing to do; all ZIPs already processed.")
        return
    with stats_lock:
        stats["total"] = len(zips)
    log(f"[JOB] Starting batch | total={len(zips)} | input={input_path} | output={output_path} | images={SAVE_IMAGES}")
    image_dir = pathlib.Path(output_path).resolve().parent / "images"
    if SAVE_IMAGES:
        image_dir.mkdir(parents=True, exist_ok=True)
    else:
        log("[INFO] Image download is OFF. Only data will be saved.")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(process_zip, z, cols, out_path, image_dir): z for z in zips}
        for _ in as_completed(futs):
            if stop_event.is_set():
                break
    close_all_browsers()


# ---------- GUI ----------
def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    app = tk.Tk()
    app.title("UPS Priority Script")
    app.geometry("700x300")
    # Simple dark styling so buttons/inputs stay visible on macOS defaults.
    BG = "#15161a"
    FG = "#f5f5f5"          # label text
    FG_SUB = "#d1d5db"
    BTN_BG = "#2f3138"
    BTN_ACTIVE = "#3c4049"
    BTN_FG = "#111111"      # dark text stays visible on macOS light buttons
    BTN_DISABLED_FG = "#777777"
    ENTRY_BG = "#1f2127"
    ENTRY_FG = "#f5f5f5"
    BORDER = "#3f434a"
    HL = "#4f46e5"
    app.configure(bg=BG)
    app.option_add("*Font", "Helvetica 11")
    app.option_add("*Button.Font", "Helvetica 11 bold")

    input_var = tk.StringVar()
    output_var = tk.StringVar(value=str(pathlib.Path("ups_output.csv").resolve()))
    status_var = tk.StringVar(value="Processed: 0 | Skipped: 0")

    def choose_input():
        path = filedialog.askopenfilename(
            title="Select input file",
            filetypes=[("CSV/XLSX/TXT", "*.csv;*.xlsx;*.xlsm;*.xls;*.txt"), ("All files", "*.*")],
        )
        if path:
            input_var.set(path)
            if not output_var.get():
                output_var.set(str(pathlib.Path(path).with_suffix(".csv")))

    def choose_output():
        path = filedialog.asksaveasfilename(
            title="Select output CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
        )
        if path:
            output_var.set(path)

    def refresh_status():
        with stats_lock:
            total = stats.get("total", 0)
            proc = stats.get("processed", 0)
            skip = stats.get("skipped", 0)
            total_txt = f"/{total}" if total else ""
            status_var.set(f"Processed: {proc}{total_txt} | Skipped: {skip}")
        app.after(500, refresh_status)

    worker_thread = None

    def start():
        nonlocal worker_thread
        if worker_thread and worker_thread.is_alive():
            messagebox.showinfo("UPS", "Job already running.")
            return
        stop_event.clear()
        pause_event.set()
        with stats_lock:
            stats["processed"] = 0
            stats["skipped"] = 0
        inp = input_var.get().strip()
        outp = output_var.get().strip()
        if not inp:
            messagebox.showerror("UPS", "Please select an input file.")
            return
        if not outp:
            messagebox.showerror("UPS", "Please select an output file.")
            return

        def run_job():
            try:
                run_batch(inp, outp)
                messagebox.showinfo("UPS", f"Done. Output: {outp}")
            except Exception as e:
                messagebox.showerror("UPS", f"Error: {e}")
                log(f"[ERROR] {e}")

        worker_thread = threading.Thread(target=run_job, daemon=True)
        worker_thread.start()

    def pause():
        pause_event.clear()

    def resume():
        pause_event.set()

    def on_close():
        stop_event.set()
        pause_event.set()
        close_all_browsers()
        app.destroy()

    # Layout frame for padding
    main = tk.Frame(app, bg=BG)
    main.grid(row=0, column=0, sticky="nsew", padx=16, pady=16)
    main.columnconfigure(1, weight=1)

    tk.Label(main, text="Input File:", bg=BG, fg=FG, font=("Helvetica", 11, "bold")).grid(
        row=0, column=0, sticky="w", padx=(0, 10), pady=8
    )
    tk.Entry(
        main,
        textvariable=input_var,
        width=55,
        bg=ENTRY_BG,
        fg=ENTRY_FG,
        insertbackground=ENTRY_FG,
        relief="flat",
        highlightthickness=1,
        highlightbackground=BORDER,
    ).grid(row=0, column=1, padx=(0, 10), sticky="ew")
    tk.Button(
        main,
        text="Browse",
        command=choose_input,
        bg=BTN_BG,
        fg=BTN_FG,
        activebackground=BTN_ACTIVE,
        activeforeground=BTN_FG,
        disabledforeground=BTN_DISABLED_FG,
        relief="flat",
        highlightthickness=0,
        padx=14,
    ).grid(row=0, column=2, padx=(0, 0))

    tk.Label(main, text="Output CSV:", bg=BG, fg=FG, font=("Helvetica", 11, "bold")).grid(
        row=1, column=0, sticky="w", padx=(0, 10), pady=8
    )
    tk.Entry(
        main,
        textvariable=output_var,
        width=55,
        bg=ENTRY_BG,
        fg=ENTRY_FG,
        insertbackground=ENTRY_FG,
        relief="flat",
        highlightthickness=1,
        highlightbackground=BORDER,
    ).grid(row=1, column=1, padx=(0, 10), sticky="ew")
    tk.Button(
        main,
        text="Browse",
        command=choose_output,
        bg=BTN_BG,
        fg=BTN_FG,
        activebackground=BTN_ACTIVE,
        activeforeground=BTN_FG,
        disabledforeground=BTN_DISABLED_FG,
        relief="flat",
        highlightthickness=0,
        padx=14,
    ).grid(row=1, column=2)

    # Note about leading zeros
    tk.Label(
        main,
        text="Tip: Use XLSX if ZIPs have leading zeros (00001). CSV/TXT edited in Excel may drop zeros.",
        bg=BG,
        fg=FG_SUB,
        font=("Helvetica", 10),
    ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(4, 6))

    # Action buttons row
    btn_frame = tk.Frame(main, bg=BG)
    btn_frame.grid(row=3, column=0, columnspan=3, pady=14, sticky="w")
    btn_frame.columnconfigure((0, 1, 2), weight=1, uniform="btns")

    tk.Button(
        btn_frame,
        text="Start",
        command=start,
        width=12,
        bg=HL,
        fg=BTN_FG,  # use dark text so it stays readable even if macOS forces a light button background
        activebackground="#4338ca",
        activeforeground=BTN_FG,
        disabledforeground=BTN_DISABLED_FG,
        relief="flat",
        highlightthickness=0,
        padx=4,
    ).grid(row=0, column=0, padx=6)
    tk.Button(
        btn_frame,
        text="Pause",
        command=pause,
        width=12,
        bg=BTN_BG,
        fg=BTN_FG,
        activebackground=BTN_ACTIVE,
        activeforeground=BTN_FG,
        disabledforeground=BTN_DISABLED_FG,
        relief="flat",
        highlightthickness=0,
        padx=4,
    ).grid(row=0, column=1, padx=6)
    tk.Button(
        btn_frame,
        text="Resume",
        command=resume,
        width=12,
        bg=BTN_BG,
        fg=BTN_FG,
        activebackground=BTN_ACTIVE,
        activeforeground=BTN_FG,
        disabledforeground=BTN_DISABLED_FG,
        relief="flat",
        highlightthickness=0,
        padx=4,
    ).grid(row=0, column=2, padx=6)

    status_label = tk.Label(
        main,
        textvariable=status_var,
        font=("Helvetica", 11, "bold"),
        bg=BG,
        fg=FG_SUB,
    )
    status_label.grid(row=4, column=0, columnspan=3, pady=(6, 0), sticky="w")

    refresh_status()
    app.protocol("WM_DELETE_WINDOW", on_close)
    app.mainloop()


def main():
    if len(sys.argv) >= 2:
        if sys.argv[1] == "--nogui":
            input_file = sys.argv[2] if len(sys.argv) > 2 else ""
            output_file = sys.argv[3] if len(sys.argv) > 3 else "ups_output.csv"
            if not input_file:
                print("Usage: ups_priority.py --nogui <input.csv|xlsx> [output.csv]")
                sys.exit(1)
            run_batch(input_file, output_file)
            print(f"Done. Output: {output_file}")
            return
        if sys.argv[1] == "--debug-headful":
            zip_arg = sys.argv[2] if len(sys.argv) > 2 else ""
            if not zip_arg:
                print("Usage: ups_priority.py --debug-headful <zip>")
                sys.exit(1)
            debug_headful(zip_arg)
            return
    launch_gui()


if __name__ == "__main__":
    main()
